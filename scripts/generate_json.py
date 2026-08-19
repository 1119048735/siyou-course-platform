import json
import os
from openpyxl import load_workbook


EXCEL_FILE = "data/course_manage.xlsx"
OUTPUT_DIR = "data/courses"

os.makedirs(OUTPUT_DIR, exist_ok=True)


# ============================================================
# 工具函数
# ============================================================

def clean_text(value):
    """
    清理 Excel 文本：
    - None -> ""
    - 去掉前后空格
    - 去掉前后 Tab
    - 去掉前后换行
    """
    if value is None:
        return ""

    return str(value).strip()


def normalize_course_id(value):
    """
    课程 ID 统一成 3 位：
    1 -> 001
    5 -> 005
    14 -> 014
    """
    if value is None:
        return ""

    # Excel 有时可能读成 1.0
    if isinstance(value, float) and value.is_integer():
        value = int(value)

    return str(value).strip().zfill(3)


def normalize_id(value):
    """
    lesson_id / chapter_id：
    如果 Excel 读取成 1.0，则转成 1
    """
    if value is None:
        return None

    if isinstance(value, float) and value.is_integer():
        return int(value)

    return value


# ============================================================
# 打开 Excel
# ============================================================

wb = load_workbook(
    EXCEL_FILE,
    data_only=True
)


required_sheets = [
    "courses",
    "lessons",
    "chapters"
]


for sheet_name in required_sheets:

    if sheet_name not in wb.sheetnames:

        raise ValueError(
            f"缺少工作表：{sheet_name}"
        )


courses_sheet = wb["courses"]
lessons_sheet = wb["lessons"]
chapters_sheet = wb["chapters"]


# ============================================================
# 读取课程
# ============================================================

courses = {}


for row in courses_sheet.iter_rows(
    min_row=2,
    values_only=True
):

    if not row[0]:
        continue


    cid = normalize_course_id(row[0])


    courses[cid] = {

        "course_id": cid,

        "course_name":
            clean_text(row[1])
            if len(row) > 1
            else "",

        "stage":
            clean_text(row[2])
            if len(row) > 2
            else "",

        "badge":
            clean_text(row[3])
            if len(row) > 3
            else "",

        "cover":
            clean_text(row[4])
            if len(row) > 4
            else "",

        "description":
            clean_text(row[5])
            if len(row) > 5
            else "",

        "sort":
            row[6]
            if len(row) > 6 and row[6] is not None
            else 0

    }


# ============================================================
# 读取全部 lessons
# ============================================================

lessons = {}


for row in lessons_sheet.iter_rows(
    min_row=2,
    values_only=True
):

    if not row[0]:
        continue


    cid = normalize_course_id(row[0])

    lesson_id = normalize_id(row[1])


    if lesson_id is None:
        continue


    lessons.setdefault(
        cid,
        {}
    )


    if lesson_id in lessons[cid]:

        print(
            f"警告：发现重复 lesson_id："
            f"课程 {cid}，lesson_id {lesson_id}"
        )


    lessons[cid][lesson_id] = {

        "lesson_id": lesson_id,

        "title":
            clean_text(row[2])
            if len(row) > 2
            else "",

        "video_file":
            clean_text(row[3])
            if len(row) > 3
            else "",

        "video_url":
            clean_text(row[4])
            if len(row) > 4
            else ""

    }


# ============================================================
# 读取章节
# ============================================================

chapters = {}


for row in chapters_sheet.iter_rows(
    min_row=2,
    values_only=True
):

    if not row[0]:
        continue


    cid = normalize_course_id(row[0])

    chapter_id = normalize_id(row[1])

    chapter_name = (
        clean_text(row[2])
        if len(row) > 2
        else ""
    )

    lesson_id = (
        normalize_id(row[3])
        if len(row) > 3
        else None
    )


    # --------------------------------------------
    # 没有章节的课程：
    # chapter_id 和 chapter_name 都为空
    # 直接跳过，不创建空章节
    # --------------------------------------------

    if chapter_id is None and not chapter_name:
        continue


    if chapter_id is None:
        print(
            f"警告：课程 {cid} "
            f"存在 chapter_name，"
            f"但 chapter_id 为空"
        )

        continue


    chapter_key = str(chapter_id)


    chapters.setdefault(
        cid,
        {}
    )


    # --------------------------------------------
    # 第一次遇到这个章节
    # --------------------------------------------

    if chapter_key not in chapters[cid]:

        chapters[cid][chapter_key] = {

            "chapter_id": chapter_key,

            "chapter_name": chapter_name,

            "lessons": []

        }

    else:

        # 检查同一个 chapter_id
        # 是否出现不同 chapter_name

        old_name = chapters[cid][chapter_key][
            "chapter_name"
        ]

        if (
            chapter_name
            and old_name
            and chapter_name != old_name
        ):

            print(
                f"警告：课程 {cid} "
                f"chapter_id {chapter_key} "
                f"出现不同章节名："
                f"「{old_name}」 / "
                f"「{chapter_name}」"
            )


    # --------------------------------------------
    # 从 lessons 表获取完整课节信息
    # --------------------------------------------

    if lesson_id is None:
        continue


    lesson = lessons.get(
        cid,
        {}
    ).get(
        lesson_id
    )


    if lesson:

        chapters[cid][chapter_key][
            "lessons"
        ].append(
            lesson
        )

    else:

        print(
            f"警告：章节表找不到对应课节："
            f"课程 {cid}，"
            f"lesson_id {lesson_id}"
        )


# ============================================================
# 数据检查
# ============================================================

print()
print("==============================")
print("开始检查课程数据")
print("==============================")


# lessons 表中有课程，
# 但是 courses 表没有课程基础信息

for cid in lessons:

    if cid not in courses:

        print(
            f"错误：课程 {cid} "
            f"存在 lessons，"
            f"但 courses 表没有该课程"
        )


# courses 表有课程，
# 但 lessons 表没有课节

for cid in courses:

    if cid not in lessons:

        print(
            f"警告：课程 {cid} "
            f"没有任何 lesson"
        )


# 检查视频地址

for cid, course_lessons in lessons.items():

    for lesson_id, lesson in course_lessons.items():

        if not lesson["video_url"]:

            print(
                f"警告：课程 {cid} "
                f"第 {lesson_id} 节 "
                f"没有 video_url"
            )


print("==============================")
print("数据检查完成")
print("==============================")
print()


# ============================================================
# 生成 JSON
# ============================================================

generated_count = 0


for cid, course in sorted(
    courses.items()
):

    data = course.copy()


    # --------------------------------------------
    # 有章节课程
    # --------------------------------------------

    if (
        cid in chapters
        and chapters[cid]
    ):

        data["chapters"] = list(
            chapters[cid].values()
        )


        # 按 chapter_id 排序
        try:

            data["chapters"].sort(
                key=lambda x: int(
                    x["chapter_id"]
                )
            )

        except (ValueError, TypeError):

            pass


        # 每章课节按 lesson_id 排序
        for chapter in data["chapters"]:

            try:

                chapter["lessons"].sort(
                    key=lambda x:
                    int(x["lesson_id"])
                )

            except (ValueError, TypeError):

                pass


        data["total_lessons"] = sum(

            len(chapter["lessons"])

            for chapter
            in data["chapters"]

        )


    # --------------------------------------------
    # 无章节课程
    # --------------------------------------------

    else:

        data["lessons"] = list(

            lessons.get(
                cid,
                {}
            ).values()

        )


        try:

            data["lessons"].sort(
                key=lambda x:
                int(x["lesson_id"])
            )

        except (ValueError, TypeError):

            pass


        data["total_lessons"] = len(
            data["lessons"]
        )


    # --------------------------------------------
    # 写入 JSON
    # --------------------------------------------

    output_file = os.path.join(
        OUTPUT_DIR,
        f"{cid}.json"
    )


    with open(
        output_file,
        "w",
        encoding="utf-8"
    ) as f:

        json.dump(
            data,
            f,
            ensure_ascii=False,
            indent=2
        )


    generated_count += 1


    print(
        f"生成完成：{cid} "
        f"{data['course_name']} "
        f"（{data['total_lessons']} 节）"
    )


# ============================================================
# 最终结果
# ============================================================

print()
print("==============================")
print(
    f"全部完成，共生成 "
    f"{generated_count} 套课程"
)
print("==============================")
