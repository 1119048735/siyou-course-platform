import os
import re
import shutil
import sqlite3
from datetime import datetime

from openpyxl import load_workbook


# ============================================================
# 配置
# ============================================================

EXCEL_FILE = "/home/ubuntu/users_manage.xlsx"

DB_FILE = "/var/lib/siyou-course-platform/devices.db"

BACKUP_DIR = "/var/lib/siyou-course-platform/backups"


# ============================================================
# 手机号清理
# ============================================================

def clean_phone(value):

    if value is None:
        return ""


    # Excel 如果把手机号识别成数字
    if isinstance(value, float):

        if value.is_integer():
            value = int(value)


    phone = str(value).strip()


    # 去掉空格、横线等非数字字符
    phone = re.sub(
        r"\D",
        "",
        phone
    )


    return phone


# ============================================================
# 验证手机号
# ============================================================

def is_valid_phone(phone):

    return bool(
        re.fullmatch(
            r"1\d{10}",
            phone
        )
    )


# ============================================================
# 读取 Excel
# ============================================================

def load_phones():

    if not os.path.exists(
        EXCEL_FILE
    ):

        raise FileNotFoundError(
            f"找不到手机号表格：{EXCEL_FILE}"
        )


    wb = load_workbook(
        EXCEL_FILE,
        data_only=True
    )


    ws = wb.active


    phones = []

    invalid_rows = []


    for row_number, row in enumerate(
        ws.iter_rows(
            min_row=1,
            values_only=True
        ),
        start=1
    ):

        if not row:
            continue


        value = row[0]


        if value is None:
            continue


        raw_text = str(value).strip()


        # 自动跳过表头
        if (
            row_number == 1
            and raw_text.lower()
            in [
                "phone",
                "手机号",
                "手机号码"
            ]
        ):

            continue


        phone = clean_phone(
            value
        )


        if not phone:
            continue


        if not is_valid_phone(
            phone
        ):

            invalid_rows.append(
                (
                    row_number,
                    raw_text
                )
            )

            continue


        phones.append(
            phone
        )


    # 自动去重
    phones = sorted(
        set(phones)
    )


    return (
        phones,
        invalid_rows
    )


# ============================================================
# 创建数据库表
# ============================================================

def ensure_tables(conn):

    cursor = conn.cursor()


    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            phone TEXT PRIMARY KEY,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """
    )


    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS devices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            phone TEXT NOT NULL,
            device_id TEXT NOT NULL,
            first_login TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            last_login TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(phone, device_id)
        )
        """
    )


    cursor.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_devices_phone
        ON devices(phone)
        """
    )


    conn.commit()


# ============================================================
# 数据库备份
# ============================================================

def backup_database():

    os.makedirs(
        BACKUP_DIR,
        exist_ok=True
    )


    if not os.path.exists(
        DB_FILE
    ):

        return None


    timestamp = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )


    backup_file = os.path.join(

        BACKUP_DIR,

        f"devices_{timestamp}.db"

    )


    shutil.copy2(

        DB_FILE,

        backup_file

    )


    return backup_file


# ============================================================
# 同步用户
# ============================================================

def sync_users(
    conn,
    excel_phones
):

    cursor = conn.cursor()


    # 当前数据库手机号
    cursor.execute(
        """
        SELECT phone
        FROM users
        """
    )


    db_phones = {
        row[0]
        for row
        in cursor.fetchall()
    }


    excel_set = set(
        excel_phones
    )


    # Excel新增的手机号
    to_add = sorted(
        excel_set
        -
        db_phones
    )


    # Excel删除的手机号
    to_remove = sorted(
        db_phones
        -
        excel_set
    )


    # ========================================================
    # 新增用户
    # ========================================================

    for phone in to_add:

        cursor.execute(
            """
            INSERT OR IGNORE INTO users (
                phone
            )
            VALUES (?)
            """,
            (
                phone,
            )
        )


    # ========================================================
    # 删除用户
    # 同时删除他的设备绑定
    # ========================================================

    for phone in to_remove:

        cursor.execute(
            """
            DELETE FROM devices
            WHERE phone = ?
            """,
            (
                phone,
            )
        )


        cursor.execute(
            """
            DELETE FROM users
            WHERE phone = ?
            """,
            (
                phone,
            )
        )


    conn.commit()


    return (
        to_add,
        to_remove
    )


# ============================================================
# 主程序
# ============================================================

def main():

    print()
    print(
        "===================================="
    )
    print(
        "开始同步用户手机号"
    )
    print(
        "===================================="
    )
    print()


    phones, invalid_rows = (
        load_phones()
    )


    # ========================================================
    # Excel格式检查
    # ========================================================

    if invalid_rows:

        print(
            "发现格式错误的手机号："
        )

        print()


        for (
            row_number,
            value
        ) in invalid_rows:

            print(
                f"第 {row_number} 行：{value}"
            )


        print()
        print(
            "请先修正 Excel 后再重新运行。"
        )

        raise SystemExit(1)


    if not phones:

        print(
            "错误：Excel 中没有找到任何有效手机号。"
        )

        print(
            "为避免误删全部用户，本次同步已取消。"
        )

        raise SystemExit(1)


    print(
        f"Excel 有效手机号：{len(phones)}"
    )


    # ========================================================
    # 打开数据库
    # ========================================================

    os.makedirs(
        os.path.dirname(
            DB_FILE
        ),
        exist_ok=True
    )


    conn = sqlite3.connect(
        DB_FILE
    )


    try:

        ensure_tables(
            conn
        )


        # ====================================================
        # 获取同步前数据
        # ====================================================

        cursor = conn.cursor()


        cursor.execute(
            """
            SELECT COUNT(*)
            FROM users
            """
        )


        before_count = (
            cursor.fetchone()[0]
        )


        print(
            f"数据库当前用户：{before_count}"
        )


        # ====================================================
        # 先备份
        # ====================================================

        backup_file = (
            backup_database()
        )


        if backup_file:

            print(
                f"数据库备份完成：{backup_file}"
            )


        # ====================================================
        # 同步
        # ====================================================

        (
            to_add,
            to_remove
        ) = sync_users(
            conn,
            phones
        )


        # ====================================================
        # 同步后统计
        # ====================================================

        cursor.execute(
            """
            SELECT COUNT(*)
            FROM users
            """
        )


        after_count = (
            cursor.fetchone()[0]
        )


        print()
        print(
            "===================================="
        )

        print(
            f"新增用户：{len(to_add)}"
        )

        print(
            f"删除用户：{len(to_remove)}"
        )

        print(
            f"最终用户：{after_count}"
        )

        print(
            "===================================="
        )


        # ====================================================
        # 显示新增手机号
        # ====================================================

        if to_add:

            print()
            print(
                "新增手机号："
            )


            for phone in to_add:

                print(
                    phone
                )


        # ====================================================
        # 显示删除手机号
        # ====================================================

        if to_remove:

            print()
            print(
                "删除手机号："
            )


            for phone in to_remove:

                print(
                    phone
                )


        print()
        print(
            "用户手机号同步完成。"
        )
        print()


    finally:

        conn.close()


if __name__ == "__main__":

    main()
